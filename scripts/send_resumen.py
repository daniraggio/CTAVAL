"""
send_resumen.py
Toma screenshot del dashboard, lo sube a GitHub y manda el mail via Gmail SMTP.
Ruta en el repo: scripts/send_resumen.py
"""

import os, base64, requests, json, smtplib, hashlib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from datetime import datetime, timezone, timedelta
ARG_TZ = timezone(timedelta(hours=-3))
from pathlib import Path
from playwright.sync_api import sync_playwright

GMAIL_USER      = "jarvis.aconcagua@gmail.com"
GMAIL_APP_PASS  = os.environ.get("GMAIL_APP_PASSWORD", "")
FROM_EMAIL      = "jarvis.aconcagua@gmail.com"
TO_EMAILS       = ["draggio@aconcaguaenergia.com"]

GH_TOKEN        = os.environ.get("GITHUB_TOKEN", "")
REPO            = "daniraggio/CTAVAL"
DASHBOARD_URL   = "https://daniraggio.github.io/CTAVAL/"
DASHBOARD_LINK  = "http://10.203.16.33/ctavav/index.html"
SCREENSHOT_PATH = Path("/tmp/resumen.png")
IMG_REPO_PATH   = "screenshots/resumen_latest.png"
IMG_PUBLIC_URL  = f"https://raw.githubusercontent.com/{REPO}/main/{IMG_REPO_PATH}"
STATE_REPO_PATH = "data/.last_email_state.json"
DATA_DIR        = Path(__file__).resolve().parent.parent / "data"

MESES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
         "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]


def take_screenshot():
    print("  → Abriendo dashboard...")
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page(viewport={"width": 1280, "height": 900})
        page.goto(DASHBOARD_URL, timeout=60_000)
        page.wait_for_load_state("networkidle")
        page.wait_for_timeout(6_000)
        try:
            # Select the last month that has actual data (may differ from current calendar month)
            last_month_with_data = page.evaluate("""() => {
                const sel = document.getElementById('monthSel');
                if (!sel) return null;
                // Find last month with energy data
                const opts = Array.from(sel.options).map(o => o.value);
                for (let i = opts.length - 1; i >= 0; i--) {
                    const k = opts[i];
                    const eRows = ALL_DATA[k] && ALL_DATA[k]['AVALTG23-ENERGIA'];
                    if (eRows && eRows.rows && eRows.rows.length > 0) return k;
                }
                return opts[opts.length - 1];
            }""")
            if last_month_with_data:
                page.evaluate(f"selectMonth('{last_month_with_data}')")
                page.wait_for_timeout(2_000)
            # Navigate to Resumen (USD) section
            page.evaluate("showSec('resumenUSD')")
            page.wait_for_timeout(2_000)
            page.locator("#sec-resumenUSD").screenshot(path=str(SCREENSHOT_PATH))
        except:
            page.screenshot(path=str(SCREENSHOT_PATH))
        browser.close()
    print(f"  → Screenshot: {SCREENSHOT_PATH.stat().st_size // 1024} KB")


def upload_screenshot_to_github():
    print("  → Subiendo screenshot a GitHub...")
    hdrs = {
        "Authorization": f"Bearer {GH_TOKEN}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28"
    }
    content = base64.b64encode(SCREENSHOT_PATH.read_bytes()).decode()

    sha = None
    meta = requests.get(f"https://api.github.com/repos/{REPO}/contents/{IMG_REPO_PATH}", headers=hdrs)
    if meta.ok:
        sha = meta.json().get("sha")

    body = {"message": "screenshot update", "content": content, "branch": "main"}
    if sha:
        body["sha"] = sha

    resp = requests.put(
        f"https://api.github.com/repos/{REPO}/contents/{IMG_REPO_PATH}",
        headers=hdrs, json=body, timeout=60
    )
    if not resp.ok:
        raise Exception(f"GitHub upload error: {resp.text}")
    print(f"  → Imagen disponible en: {IMG_PUBLIC_URL}")


def build_html_email():
    now = datetime.now(ARG_TZ)
    mes = MESES[now.month-1]
    fecha_str = now.strftime("%d/%m/%Y %H:%M")
    img_url = f"{IMG_PUBLIC_URL}?t={int(now.timestamp())}"

    html = f"""<!DOCTYPE html>
<html>
<head><meta charset="utf-8"></head>
<body style="margin:0;padding:0;background:#0d1117;font-family:system-ui,sans-serif">
  <div style="max-width:700px;margin:0 auto;padding:32px 16px">
    <div style="margin-bottom:24px">
      <div style="font-size:22px;font-weight:700;color:#e6edf3">Central Térmica Alto Valle</div>
      <div style="font-size:14px;color:#8b949e;margin-top:4px">Actualización diaria — {fecha_str}</div>
    </div>
    <div style="background:#161b22;border:1px solid #30363d;border-radius:8px;padding:20px;margin-bottom:20px">
      <div style="font-size:13px;font-weight:600;color:#8b949e;text-transform:uppercase;letter-spacing:.06em;margin-bottom:12px">Resumen — {mes} {now.year}</div>
      <img src="{img_url}" width="640" style="width:100%;border-radius:6px;display:block" alt="Resumen del mes"/>
    </div>
    <div style="text-align:center;margin:28px 0">
      <a href="{DASHBOARD_LINK}" style="background:#238636;color:#ffffff;text-decoration:none;padding:12px 28px;border-radius:6px;font-weight:600;font-size:14px;display:inline-block">
        Ver Dashboard Completo →
      </a>
    </div>
    <div style="font-size:11px;color:#484f58;text-align:center;margin-top:24px">
      Este mail se genera automáticamente todos los días.
    </div>
  </div>
</body>
</html>"""
    subject = f"Reporte Central Térmica Alto Valle — {now.day} de {mes} {now.year}"
    return html, subject


def send_email(html_body, subject):
    msg = MIMEMultipart('alternative')
    msg['Subject'] = subject
    msg['From']    = FROM_EMAIL
    msg['To']      = ', '.join(TO_EMAILS)
    msg.attach(MIMEText(html_body, 'html'))

    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(GMAIL_USER, GMAIL_APP_PASS)
        smtp.sendmail(FROM_EMAIL, TO_EMAILS, msg.as_string())
    print(f"  ✅ Mail enviado a: {', '.join(TO_EMAILS)}")


def current_month_xls():
    """Devuelve el Path del .xls del último mes con datos disponibles."""
    now = datetime.now(ARG_TZ)
    # Try current month first, then go back until we find an existing file
    for delta in range(3):
        month = now.month - delta
        year = now.year
        if month <= 0:
            month += 12
            year -= 1
        path = DATA_DIR / f"AVAL_{year}_{month:02d}.xls"
        if path.exists():
            return path
    return DATA_DIR / f"AVAL_{now.year}_{now.month:02d}.xls"


def file_hash(path):
    """Hash solo las filas de datos del XLS (ignora metadatos como fecha de generación)."""
    if not path.exists():
        return None
    try:
        import openpyxl
        from datetime import datetime as dt
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        h = hashlib.sha256()
        # Hash the data rows of key sheets only
        key_sheets = ['AVALTG23-ENERGIA', 'AVALTG22-ENERGIA', 'AVALTV11-ENERGIA', 'AVALTV12-ENERGIA',
                      'PRECIOS-CMO', 'AVALTG23-CVPCOMBREM-GX', 'AVALTG22-CVPCOMBREM-GX']
        for sheet in key_sheets:
            if sheet not in wb.sheetnames:
                continue
            for row in wb[sheet].iter_rows(values_only=True):
                if isinstance(row[0], dt):
                    # Only hash date + numeric values, skip metadata rows
                    row_str = str(row[0].date()) + '|' + '|'.join(
                        str(round(v, 4)) if isinstance(v, float) else str(v)
                        for v in row[2:26]
                    )
                    h.update(row_str.encode())
        wb.close()
        return h.hexdigest()
    except Exception:
        # Fallback to full file hash
        return hashlib.sha256(path.read_bytes()).hexdigest()


def get_last_email_state():
    hdrs = {"Authorization": f"Bearer {GH_TOKEN}", "Accept": "application/vnd.github+json"}
    resp = requests.get(f"https://api.github.com/repos/{REPO}/contents/{STATE_REPO_PATH}", headers=hdrs, timeout=30)
    if not resp.ok:
        return {}, None
    data = resp.json()
    sha = data.get("sha")
    try:
        content = base64.b64decode(data["content"]).decode()
        return json.loads(content), sha
    except Exception:
        return {}, sha


def save_email_state(state, sha):
    hdrs = {"Authorization": f"Bearer {GH_TOKEN}", "Accept": "application/vnd.github+json"}
    content = base64.b64encode(json.dumps(state, indent=2).encode()).decode()
    body = {"message": "update last_email_state", "content": content, "branch": "main"}
    if sha:
        body["sha"] = sha
    resp = requests.put(f"https://api.github.com/repos/{REPO}/contents/{STATE_REPO_PATH}", headers=hdrs, json=body, timeout=30)
    if not resp.ok:
        raise Exception(f"GitHub state update error: {resp.text}")


if __name__ == "__main__":
    now = datetime.now(ARG_TZ)
    print(f"[{now:%H:%M:%S}] Generando resumen...")
    try:
        xls_path = current_month_xls()
        current_hash = file_hash(xls_path)
        state, state_sha = get_last_email_state()

        # If state uses old hash method (no 'v' key), force resend once to recalibrate
        state_hash = state.get("hash") if state.get("v") == 2 else None

        if current_hash and state_hash == current_hash:
            print(f"  ℹ️  Sin novedades en {xls_path.name} desde el último mail — no se envía.")
        else:
            take_screenshot()
            upload_screenshot_to_github()
            html, subject = build_html_email()
            print(f"  → Enviando a: {', '.join(TO_EMAILS)}")
            send_email(html, subject)
            if current_hash:
                state["hash"] = current_hash
                state["fecha"] = now.isoformat()
                state["v"] = 2  # mark as data-only hash
                state["xls"] = xls_path.name
                save_email_state(state, state_sha)
    except Exception as e:
        print(f"  ❌ Error: {e}")
        raise
